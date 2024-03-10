from datetime import datetime, timedelta
from env import KOMAR, VANIN
from dateutil.relativedelta import relativedelta
from time import sleep
from functions import connect, cursor
import openpyxl as op
import requests
import time


# def compare_dates():
#     cursor.execute(f'SELECT verification_date, valid_date FROM uploaded_data WHERE si_number = "I3223557"')
#     fetch = cursor.fetchone()
#     date1 = datetime.strptime(fetch[0][:10], '%Y-%m-%d')
#     date2 = datetime.strptime(fetch[1][:10], '%Y-%m-%d')
#     if date1.day == date2.day:
#         print(f'{date1.date()}\n'
#               f'{date2.date()}')
#         print('Found')
#     else:
#         print(date1.date(), date2.date())
#         print('Not found')


def strip_numbers():
    cursor.execute('SELECT si_number FROM uploaded_data WHERE si_number LIKE " %" or si_number LIKE "% "')
    fetch = cursor.fetchall()
    for counter in fetch:
        striped_number = counter[0].strip()
        cursor.execute("""
                            UPDATE uploaded_data
                            SET si_number = ?
                            WHERE si_number = ?;
                        """, (striped_number, counter[0]))
        connect.commit()


def get_by_date(date, intern):
    query = "SELECT count(*) FROM uploaded_data WHERE strftime('%s', verification_date) = strftime('%s', ?)  AND intern = ?"
    cursor.execute(query, (date, intern))
    fetch = cursor.fetchone()
    return fetch


def count_verifies(date_1, date_2, verifier: str):
    current_date = date_1
    result = {}
    while current_date < date_2:
        count = get_by_date(current_date, verifier.upper())
        result[datetime.strftime(current_date, '%d.%m.%Y')] = count[0]
        current_date += timedelta(days=1)
    return result


def generate_excel(verifies: dict, intern: str):
    wb = op.Workbook()
    sheet = wb.active
    i = 1
    for date, count in verifies.items():
        date_cell = sheet.cell(row = 2, column = i)
        date_cell.value = f"{date}"
        count_cell = sheet.cell(row = 3, column = i)
        if count:
            count_cell.value = count
        i += 1
    sum_cell = sheet.cell(row = 3, column = i)
    sum_cell.value = sum(verifies.values())
    wb.save(f'{intern}.xlsx')


def get_by_date_2(date, standard):
    query = "SELECT count(*) FROM uploaded_data WHERE strftime('%s', verification_date) = strftime('%s', ?) AND standart = ?"
    cursor.execute(query, (date,standard))
    fetch = cursor.fetchone()
    return fetch


def get_standards():
    query = "SELECT standart_manufacture_num FROM standarts"
    cursor.execute(query)
    fetch = cursor.fetchall()
    result = set([i[0] for i in fetch])
    return list(result)


def count_verifies_2(date_1, date_2, standards):
    current_date = date_1
    result = {}
    while current_date < date_2:
        result[datetime.strftime(current_date, '%d.%m.%Y')] = {}
        for standart in standards:
            count = get_by_date_2(current_date, standart)
            if count[0]:
                result[datetime.strftime(current_date, '%d.%m.%Y')].update({standart: count[0]})
        result[datetime.strftime(current_date, '%d.%m.%Y')] = dict(sorted(result[datetime.strftime(current_date, '%d.%m.%Y')].items(), key=lambda x: x[1], reverse=True))
        current_date += timedelta(days=1)
    return result


def get_verifies_without_vri_id(date):

    query = "SELECT ngr, si_number, verification_date FROM uploaded_data WHERE strftime('%s', verification_date) = strftime('%s', ?) AND vri_id is NULL"
    cursor.execute(query, (date, ))
    fetch = cursor.fetchall()
    return fetch if len(fetch) else []


def get_data_from_arshin(verifies_data):
    result_dict = {}
    if verifies_data:
        for data in verifies_data:
            params = {
                'mit_number': data[0],
                'mi_number': data[1],
                'verification_date': data[2],
                'year': data[2][0:4],
                'org_title': 'Индивидуальный предприниматель Дьяченко Алексей Олегович', #"ООО \"ВОДОРЕСУРС\"" # 'Индивидуальный предприниматель Дьяченко Алексей Олегович'
            }
            url = f'https://fgis.gost.ru/fundmetrology/eapi/vri'
            req = requests.get(url, params=params)

            try:
                req_json = req.json()
                if req_json['result']['count'] != 0:
                    result_dict.update({
                        data[1]: {
                            'result_docnum': req_json['result']['items'][0]['result_docnum'],
                            'vri_id': req_json['result']['items'][0]['vri_id'],
                        }
                    })
            except Exception as e:
                print(e)
                print(params)
                print(req_json)
            sleep(0.5)
    return result_dict


def update_vri_id(key, input_data) -> None:
    query = "UPDATE uploaded_data SET vri_id=?, result_docnum=? WHERE si_number=?"
    cursor.execute(query, (input_data['vri_id'], input_data['result_docnum'], key))
    connect.commit()


def set_vri_id(start_date: str, end_date: str) -> str:
    start_date = datetime.strptime(start_date, '%d.%m.%Y')
    end_date = datetime.strptime(end_date, '%d.%m.%Y')
    current_date = start_date
    while current_date < end_date:
        current_date = current_date + timedelta(days=1)
        from_db = get_verifies_without_vri_id(current_date)
        from_arshin = get_data_from_arshin(from_db)
        for k, v in from_arshin.items():
            update_vri_id(k, v)
    return 'Job is done'


    # while current_date < date_2:
    #     result[datetime.strftime(current_date, '%d.%m.%Y')] = {}
    #     for standart in standards:
    #         count = get_by_date_2(current_date, standart)
    #         if count[0]:
    #             result[datetime.strftime(current_date, '%d.%m.%Y')].update({standart: count[0]})
    #     result[datetime.strftime(current_date, '%d.%m.%Y')] = dict(
    #         sorted(result[datetime.strftime(current_date, '%d.%m.%Y')].items(), key=lambda x: x[1], reverse=True))
    #     current_date += timedelta(days=1)
    # return result


if __name__ == '__main__':
    start = datetime.now()
    set_vri_id('20.01.2024', '31.03.2024')
    end = datetime.now()
    print(f'Finished in {end - start}')

    # input_1 = input('Введите начальную дату\n')
    # start_date = datetime.strptime(input_1, '%d.%m.%Y')
    # input_2 = input('Введите конечную дату\n')
    # end_date = datetime.strptime(input_2, '%d.%m.%Y')

    # Расчёт поверок на эталон
    # standards = get_standards()
    # total = count_verifies_2(start_date, end_date, standards)
    # for date, standarts in total.items():
    #     print(date, end='\n')
    #     for standart, count in standarts.items():
    #         print(standart, count)
    #     print('\n')

    # Отчёт поверок Комарских
    # for verifier in KOMAR:
    #     verifies_count = count_verifies(start_date, end_date, verifier)
    #     if sum(verifies_count.values()) != 0:
    #         generate_excel(verifies_count, verifier)

